#!/usr/bin/env python3
"""
Генерация входных файлов для эталонных прототипов заданий 3, 9, 10, 17, 18, 22, 24, 26, 27.
По ТЗ: структура, форматы, диапазоны как у ФИПИ.
После генерации решает задачи и возвращает ответы для вписывания в эталоны.

Запуск:
  python scripts/generate_reference_attachments.py [--dry-run]
  --dry-run: только показать ответы, файлы не создавать
"""
import os
import random
import csv
from datetime import datetime, timedelta
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
PROTOTYPES = REPO_ROOT / "data" / "reference_prototypes"


def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


# ---------- Задание 3: Базы данных (.xlsx, 3 листа) ----------
def generate_task_03(base_dir: Path, level: str, seed: int = 42) -> int:
    import openpyxl
    from openpyxl.styles import Font

    random.seed(seed)
    att = ensure_dir(base_dir / "attachments")

    # Справочники
    otdely = ["Бакалея", "Молочный", "Овощи", "Хлеб", "Напитки"]
    tovary = [
        ("Крупа Гречневая", "кг", 1),
        ("Молоко 3.2%", "л", 1),
        ("Кефир 1%", "л", 1),
        ("Хлеб белый", "шт", 1),
        ("Масло сливочное", "кг", 1),
        ("Сахар", "кг", 1),
        ("Рис", "кг", 1),
    ]
    rayony = ["Заречный", "Октябрьский", "Центральный", "Северный", "Южный"]
    artikuly = list(range(1, len(tovary) + 1))  # 1..7
    id_magazinov = list(range(1, 6))  # 1..5

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Лист "Товар"
    ws_t = wb.create_sheet("Товар", 0)
    ws_t.append(["Артикул", "Отдел", "Наименование", "Единица измерения", "Количество в упаковке"])
    for i, (name, ed, kol) in enumerate(tovary, 1):
        ws_t.append([i, otdely[i % len(otdely)], name, ed, kol])

    # Лист "Магазин"
    ws_m = wb.create_sheet("Магазин", 1)
    ws_m.append(["ID магазина", "Район", "Адрес"])
    for mid in id_magazinov:
        ws_m.append([mid, rayony[mid % len(rayony)], f"ул. Примерная, {mid}"])

    # Лист "Движение товаров"
    ws_d = wb.create_sheet("Движение товаров", 2)
    ws_d.append(["ID операции", "Дата", "ID магазина", "Артикул", "Количество", "Тип операции", "Цена"])
    start_date = datetime(2024, 6, 1)
    end_date = datetime(2024, 6, 10)
    n_rows = 1500 if level == "easy" else 2500 if level == "medium" else 3000
    answer_total = 0
    op_id = 1
    for _ in range(n_rows):
        d = start_date + timedelta(days=random.randint(0, 9))
        mid = random.choice(id_magazinov)
        art = random.choice(artikuly)
        kol = random.randint(1, 20)
        tip = random.choice(["Поступление", "Продажа", "Продажа", "Продажа"])
        cena = random.randint(30, 500)
        ws_d.append([op_id, d.strftime("%d.%m.%Y"), mid, art, kol, tip, cena])
        if art == 1 and tip == "Продажа" and start_date <= d <= end_date:
            answer_total += kol * cena
        op_id += 1

    out = att / ("3_easy.xlsx" if level == "easy" else "3_001.xlsx" if level == "medium" else "3_hard.xlsx")
    wb.save(out)
    return answer_total


# ---------- Задание 9: Электронные таблицы (CSV, 3 числа в строке, пифагоровы тройки) ----------
def _pythagorean_triples(limit: int):
    triples = []
    for a in range(1, min(limit, 200)):
        for b in range(a, min(limit, 200)):
            c_sq = a * a + b * b
            c = int(c_sq ** 0.5)
            if c * c == c_sq and c <= limit:
                triples.append((a, b, c))
    return triples


def generate_task_09(base_dir: Path, level: str, seed: int = 43) -> int:
    random.seed(seed)
    att = ensure_dir(base_dir / "attachments")
    n_rows = 1500 if level == "easy" else 3000 if level == "medium" else 5000
    triples = _pythagorean_triples(100000)
    out_name = "9_easy.csv" if level == "easy" else "9.txt" if level == "medium" else "9_hard.txt"
    out = att / out_name
    count = 0
    sep = ";"
    with open(out, "w", encoding="utf-8", newline="") as f:
        for i in range(n_rows):
            if random.random() < 0.15 and triples:
                a, b, c = random.choice(triples)
                row = [a, b, c]
                count += 1
            else:
                row = [random.randint(1, 500) for _ in range(3)]
            f.write(sep.join(map(str, row)) + "\n")
    return count


# ---------- Задание 10: Поиск в тексте (.docx) ----------
def generate_task_10(base_dir: Path, level: str, seed: int = 44) -> int:
    from docx import Document
    from docx.shared import Pt

    random.seed(seed)
    att = ensure_dir(base_dir / "attachments")
    doc = Document()
    # Генерируем абзацы с вкраплениями "князь" / "Князь"
    words_ru = "все мы не что как он она они его её их там тогда уже еще раз два три раз сказал спросил ответил думал видел пришел ушел домой город деревня жизнь день ночь".split()
    target = "князь"
    target_cap = "Князь"
    n_paragraphs = 80 if level == "easy" else 120 if level == "medium" else 180
    total_count = 0
    for _ in range(n_paragraphs):
        sentence_len = random.randint(8, 25)
        words = []
        for _ in range(sentence_len):
            if random.random() < 0.04:
                w = target_cap if random.random() < 0.3 else target
                words.append(w)
                total_count += 1
            else:
                words.append(random.choice(words_ru))
        doc.add_paragraph(" ".join(words))
    out_name = "10_dostoevsky.docx" if level == "easy" else "10_kuprin.docx" if level == "medium" else "10_hard.docx"
    doc.save(att / out_name)
    return total_count


# ---------- Задание 17: Последовательность (.txt, по числу на строку) ----------
def generate_task_17(base_dir: Path, level: str, seed: int = 45) -> tuple:
    random.seed(seed)
    att = ensure_dir(base_dir / "attachments")
    n = 6000 if level == "easy" else 8000 if level == "medium" else 10000
    nums = [random.randint(-10000, 10000) for _ in range(n)]
    out = att / ("17_easy.txt" if level == "easy" else "17_medium.txt" if level == "medium" else "17_hard.txt")
    with open(out, "w", encoding="utf-8") as f:
        for x in nums:
            f.write(str(x) + "\n")
    # Решение: пары подряд, хотя бы одно кратно 3, вывести количество и макс сумму
    cnt = 0
    best = -10**9
    for i in range(len(nums) - 1):
        a, b = nums[i], nums[i + 1]
        if (a % 3 == 0) or (b % 3 == 0):
            cnt += 1
            best = max(best, a + b)
    return (cnt, best)


# ---------- Задание 18: Робот, сетка N×N (.xlsx) ----------
def generate_task_18(base_dir: Path, level: str, seed: int = 46) -> tuple:
    import openpyxl

    random.seed(seed)
    att = ensure_dir(base_dir / "attachments")
    N = 12 if level == "easy" else 15 if level == "medium" else 20
    grid = [[random.randint(1, 100) for _ in range(N)] for _ in range(N)]
    out_name = "18_easy.xlsx" if level == "easy" else "18_walls.xlsx" if level == "medium" else "18_hard.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, row in enumerate(grid):
        for j, v in enumerate(row):
            ws.cell(row=i + 1, column=j + 1, value=v)
    wb.save(att / out_name)
    # DP: max и min сумма путь (0,0) -> (N-1,N-1) только вправо/вниз
    dp_max = [[0] * N for _ in range(N)]
    dp_min = [[0] * N for _ in range(N)]
    dp_max[0][0] = dp_min[0][0] = grid[0][0]
    for j in range(1, N):
        dp_max[0][j] = dp_max[0][j - 1] + grid[0][j]
        dp_min[0][j] = dp_min[0][j - 1] + grid[0][j]
    for i in range(1, N):
        dp_max[i][0] = dp_max[i - 1][0] + grid[i][0]
        dp_min[i][0] = dp_min[i - 1][0] + grid[i][0]
    for i in range(1, N):
        for j in range(1, N):
            dp_max[i][j] = max(dp_max[i - 1][j], dp_max[i][j - 1]) + grid[i][j]
            dp_min[i][j] = min(dp_min[i - 1][j], dp_min[i][j - 1]) + grid[i][j]
    return (dp_max[N - 1][N - 1], dp_min[N - 1][N - 1])


# ---------- Задание 22: Параллельные процессы (CSV, DAG) ----------
def generate_task_22(base_dir: Path, level: str, seed: int = 47) -> int:
    random.seed(seed)
    att = ensure_dir(base_dir / "attachments")
    n_proc = 12 if level == "easy" else 15 if level == "medium" else 20
    # DAG: процессы 1..n, зависимости только на меньшие номера
    times = [random.randint(10, 500) for _ in range(n_proc)]
    deps = []
    for i in range(1, n_proc + 1):
        # 0-2 предка из [1..i-1]
        k = random.randint(0, min(2, i - 1)) if i > 1 else 0
        if k == 0:
            deps.append([])
        else:
            pred = random.sample(range(1, i), k)
            deps.append(pred)
    out_name = "22_easy.csv" if level == "easy" else "22_medium.csv" if level == "medium" else "22_hard.csv"
    with open(att / out_name, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter=";")
        for i in range(1, n_proc + 1):
            dep_str = "0" if not deps[i - 1] else ";".join(map(str, deps[i - 1]))
            w.writerow([i, times[i - 1], dep_str])
    # Решение: время завершения = time + max(конец предков)
    end_time = [0] * (n_proc + 1)
    for i in range(1, n_proc + 1):
        start = max(end_time[j] for j in deps[i - 1]) if deps[i - 1] else 0
        end_time[i] = start + times[i - 1]
    return max(end_time[1:])


# ---------- Задание 24: Строка A,B,C, макс подряд C (.txt) ----------
def generate_task_24(base_dir: Path, level: str, seed: int = 48) -> int:
    random.seed(seed)
    att = ensure_dir(base_dir / "attachments")
    size = 100_000 if level == "easy" else 500_000 if level == "medium" else 1_000_000
    abc = ["A", "B", "C"]
    # Внедряем одну длинную цепочку C и шум
    max_c_len = random.randint(50, 200) if level == "easy" else random.randint(100, 500) if level == "medium" else random.randint(200, 800)
    pos = random.randint(0, max(0, size - max_c_len - 1000))
    arr = [random.choice(abc) for _ in range(size)]
    for i in range(pos, min(pos + max_c_len, size)):
        arr[i] = "C"
    out_name = "24_easy.txt" if level == "easy" else "24_medium.txt" if level == "medium" else "24_hard.txt"
    with open(att / out_name, "w", encoding="utf-8") as f:
        f.write("".join(arr))
    return max_c_len


# ---------- Задание 26: разновидности ----------
def generate_task_26_easy(base_dir: Path, seed: int = 49) -> int:
    random.seed(seed)
    att = ensure_dir(base_dir / "attachments")
    n = 6000
    prices = [random.randint(50, 5000) for _ in range(n)]
    with open(att / "26_easy.txt", "w", encoding="utf-8") as f:
        f.write(str(n) + "\n")
        for p in prices:
            f.write(str(p) + "\n")
    prices.sort()
    k = n // 3
    discount = sum(prices[-k:]) * 0.2 + sum(prices[:-k]) * 0.1
    return int(discount)


def generate_task_26_medium(base_dir: Path, seed: int = 50) -> int:
    random.seed(seed)
    att = ensure_dir(base_dir / "attachments")
    n = 100
    events = []
    t = 0
    for _ in range(n):
        start = t
        t += random.randint(1, 20)
        end = t
        t += 1
        events.append((start, end))
    random.shuffle(events)
    with open(att / "26_events.txt", "w", encoding="utf-8") as f:
        f.write(str(n) + "\n")
        for s, e in events:
            f.write(f"{s} {e}\n")
    events.sort(key=lambda x: x[1])
    count = 0
    last_end = -1
    for s, e in events:
        if s >= last_end:
            count += 1
            last_end = e
    return count


def generate_task_26_hard(base_dir: Path, seed: int = 51) -> tuple:
    random.seed(seed)
    att = ensure_dir(base_dir / "attachments")
    K = 5
    N = 30
    luggage = []
    for _ in range(N):
        start = random.randint(0, 500)
        luggage.append((start, start + random.randint(1, 50)))
    with open(att / "26_hard.txt", "w", encoding="utf-8") as f:
        f.write(f"{K} {N}\n")
        for s, e in luggage:
            f.write(f"{s} {e}\n")
    luggage.sort(key=lambda x: x[0])
    cells = [0] * K
    count = 0
    last_cell = 0
    for start, end in luggage:
        for j in range(K):
            if cells[j] < start:
                cells[j] = end
                count += 1
                last_cell = j + 1
                break
    return (count, last_cell)


# ---------- Задание 27 ----------
# easy: два файла A и B — пары (|i-j|>=5, сумма кратна 3) -> два ответа
# medium: один файл 27_pref.txt — макс подпоследовательность с суммой % 89 == 0
# hard: один файл 27_hard.txt — первая строка K, вторая N, далее N чисел; тройка с расстоянием >= K, макс сумма
def generate_task_27_easy(base_dir: Path, seed: int = 52) -> tuple:
    random.seed(seed)
    att = ensure_dir(base_dir / "attachments")
    n_a, n_b = 50, 80000

    def solve_pairs(n: int, nums: list) -> int:
        rems = [0] * 3
        queue = []
        count = 0
        K = 5
        for x in nums:
            if len(queue) == K:
                x_old = queue.pop(0)
                rems[x_old % 3] += 1
            target_rem = (3 - (x % 3)) % 3
            count += rems[target_rem]
            queue.append(x)
        return count

    nums_a = [random.randint(-1000, 1000) for _ in range(n_a)]
    nums_b = [random.randint(-1000, 1000) for _ in range(n_b)]
    with open(att / "27_A.txt", "w", encoding="utf-8") as f:
        f.write(str(n_a) + "\n")
        for x in nums_a:
            f.write(str(x) + "\n")
    with open(att / "27_B.txt", "w", encoding="utf-8") as f:
        f.write(str(n_b) + "\n")
        for x in nums_b:
            f.write(str(x) + "\n")
    return (solve_pairs(n_a, nums_a), solve_pairs(n_b, nums_b))


def generate_task_27_medium(base_dir: Path, seed: int = 53) -> int:
    random.seed(seed)
    att = ensure_dir(base_dir / "attachments")
    n = 200000
    nums = [random.randint(-10000, 10000) for _ in range(n)]
    with open(att / "27_pref.txt", "w", encoding="utf-8") as f:
        f.write(str(n) + "\n")
        for x in nums:
            f.write(str(x) + "\n")
    min_prefix = {r: float("inf") for r in range(89)}
    min_prefix[0] = 0
    max_sum = 0
    current_sum = 0
    for x in nums:
        current_sum += x
        rem = current_sum % 89
        if min_prefix[rem] != float("inf"):
            max_sum = max(max_sum, current_sum - min_prefix[rem])
        min_prefix[rem] = min(min_prefix[rem], current_sum)
    return max_sum


def generate_task_27_hard(base_dir: Path, seed: int = 54) -> int:
    random.seed(seed)
    att = ensure_dir(base_dir / "attachments")
    K = random.randint(5, 15)
    N = 100000
    nums = [random.randint(-1000, 1000) for _ in range(N)]
    with open(att / "27_hard.txt", "w", encoding="utf-8") as f:
        f.write(str(K) + "\n")
        f.write(str(N) + "\n")
        for x in nums:
            f.write(str(x) + "\n")
    max_1 = -10**9
    max_2 = -10**9
    ans = -10**9
    for i in range(2 * K, N):
        x_prev = nums[i - 2 * K]
        if x_prev > max_1:
            max_1 = x_prev
        pair = max_1 + nums[i - K]
        if pair > max_2:
            max_2 = pair
        current_triple = max_2 + nums[i]
        if current_triple > ans:
            ans = current_triple
    return ans


def main():
    import argparse
    import tempfile
    import shutil
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Писать во временную папку, показать ответы, затем удалить")
    args = parser.parse_args()
    root = REPO_ROOT
    if args.dry_run:
        root = Path(tempfile.mkdtemp())
    prototypes_root = root / "data" / "reference_prototypes"
    results = {}

    # Task 3
    for level in ["easy", "medium", "hard"]:
        base = prototypes_root / "task_03" / level
        ensure_dir(base / "attachments")
        results[("3", level)] = str(generate_task_03(base, level))

    # Task 9
    for level in ["easy", "medium", "hard"]:
        base = prototypes_root / "task_09" / level
        ensure_dir(base / "attachments")
        results[("9", level)] = str(generate_task_09(base, level))

    # Task 10
    for level in ["easy", "medium", "hard"]:
        base = prototypes_root / "task_10" / level
        ensure_dir(base / "attachments")
        results[("10", level)] = str(generate_task_10(base, level))

    # Task 17
    for level in ["easy", "medium", "hard"]:
        base = prototypes_root / "task_17" / level
        ensure_dir(base / "attachments")
        a = generate_task_17(base, level)
        results[("17", level)] = f"{a[0]} {a[1]}"

    # Task 18
    for level in ["easy", "medium", "hard"]:
        base = prototypes_root / "task_18" / level
        ensure_dir(base / "attachments")
        a = generate_task_18(base, level)
        results[("18", level)] = f"{a[0]} {a[1]}"

    # Task 22
    for level in ["easy", "medium", "hard"]:
        base = prototypes_root / "task_22" / level
        ensure_dir(base / "attachments")
        results[("22", level)] = str(generate_task_22(base, level))

    # Task 24
    for level in ["easy", "medium", "hard"]:
        base = prototypes_root / "task_24" / level
        ensure_dir(base / "attachments")
        results[("24", level)] = str(generate_task_24(base, level))

    # Task 26
    ensure_dir(prototypes_root / "task_26" / "easy" / "attachments")
    ensure_dir(prototypes_root / "task_26" / "medium" / "attachments")
    ensure_dir(prototypes_root / "task_26" / "hard" / "attachments")
    results[("26", "easy")] = str(generate_task_26_easy(prototypes_root / "task_26" / "easy"))
    results[("26", "medium")] = str(generate_task_26_medium(prototypes_root / "task_26" / "medium"))
    ah = generate_task_26_hard(prototypes_root / "task_26" / "hard")
    results[("26", "hard")] = f"{ah[0]} {ah[1]}"

    # Task 27
    ensure_dir(prototypes_root / "task_27" / "easy" / "attachments")
    ensure_dir(prototypes_root / "task_27" / "medium" / "attachments")
    ensure_dir(prototypes_root / "task_27" / "hard" / "attachments")
    a27e = generate_task_27_easy(prototypes_root / "task_27" / "easy")
    results[("27", "easy")] = f"{a27e[0]} {a27e[1]}"
    results[("27", "medium")] = str(generate_task_27_medium(prototypes_root / "task_27" / "medium"))
    results[("27", "hard")] = str(generate_task_27_hard(prototypes_root / "task_27" / "hard"))

    # Вывод ответов для вставки в эталоны
    print("--- Ответы для эталонов (task, level -> answer) ---")
    for (task, level), ans in sorted(results.items(), key=lambda x: (int(x[0][0]), x[0][1])):
        print(f"  {task} {level}: {ans}")
    if args.dry_run:
        shutil.rmtree(root, ignore_errors=True)
        print("(dry-run: временная папка удалена; запусти без --dry-run для записи в репозиторий)")


if __name__ == "__main__":
    main()

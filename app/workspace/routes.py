"""
Workspace blueprint: файловое мини-хранилище ученика + canvas-рисование.

Endpoints:
  /workspace/files              GET    — список файлов в workspace
  /workspace/copy-from-task     POST   — скопировать файл задания в workspace
  /workspace/upload             POST   — загрузить свой файл
  /workspace/<id>/rename        POST   — переименовать файл
  /workspace/<id>               DELETE — удалить файл
  /workspace/<id>/content       GET    — содержимое файла (text / excel-json)
  /workspace/<id>/download      GET    — скачать файл

  /api/canvas/save              POST   — сохранить штрихи холста
  /api/canvas/load              GET    — загрузить свой холст
  /api/canvas/view/<user_id>    GET    — преподаватель смотрит холст ученика
"""
from __future__ import annotations

import io
import json
import logging
import mimetypes
import os
import re
import tempfile
from urllib.parse import urlparse

import requests as http_requests
import sqlalchemy as sa
from flask import (
    Blueprint, request, jsonify, current_app, send_file, abort, url_for,
)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename

from app.limiter import limiter
from app.storage.s3_service import storage
from app.utils.relationship_scope import can_user_access_student
from core.db_models import (
    Assignment,
    db,
    AssignmentTask,
    Lesson,
    LessonTask,
    StudentWorkspaceFile,
    Submission,
    TaskCanvasDrawing,
    Tasks,
)

logger = logging.getLogger(__name__)

workspace_bp = Blueprint('workspace', __name__)

ALLOWED_EXTENSIONS = {
    'txt', 'csv', 'tsv', 'py', 'cpp', 'c', 'h', 'java', 'js',
    'json', 'xml', 'html', 'css', 'md', 'log', 'ini', 'cfg',
    'xls', 'xlsx', 'xlsm', 'ods',
    'pdf', 'doc', 'docx', 'odt', 'rtf',
    'dat', 'in', 'out', 'ans',
}

TEXT_EXTENSIONS = {
    'txt', 'csv', 'tsv', 'py', 'cpp', 'c', 'h', 'java', 'js',
    'json', 'xml', 'html', 'css', 'md', 'log', 'ini', 'cfg',
    'dat', 'in', 'out', 'ans',
}

SPREADSHEET_EXTENSIONS = {'xls', 'xlsx', 'xlsm', 'ods'}

MAX_WORKSPACE_FILE_SIZE = 10 * 1024 * 1024
MAX_FILES_PER_TASK = 20

CODEMIRROR_MODES = {
    'py': 'python', 'python': 'python',
    'cpp': 'text/x-c++src', 'c': 'text/x-csrc', 'h': 'text/x-csrc',
    'java': 'text/x-java', 'js': 'javascript', 'json': 'application/json',
    'xml': 'xml', 'html': 'htmlmixed', 'css': 'css', 'md': 'markdown',
}

_workspace_tables_ok: bool | None = None


def _is_parent_user() -> bool:
    return bool(getattr(current_user, 'is_parent', lambda: False)())


def _workspace_context_scope(task_id: int, context_type: str | None, context_id: int | None) -> dict:
    """Resolve shared student-owned workspace scope for submission/lesson contexts."""
    context_type = (context_type or 'submission').strip().lower()
    # Task Workspace uses the explicit UI name ``submission_task``.  Canvas
    # storage is shared by the whole submission task and must resolve to the
    # protected submission scope, otherwise a teacher cannot view the drawing.
    if context_type == 'submission_task':
        context_type = 'submission'
    elif context_type == 'lesson_task':
        context_type = 'lesson'
    if context_type == 'submission':
        if not context_id:
            abort(400, description='context_id required')
        submission = (
            Submission.query
            .options(
                sa.orm.joinedload(Submission.student),
                sa.orm.joinedload(Submission.assignment).joinedload(Assignment.tasks),
            )
            .filter_by(submission_id=int(context_id))
            .first_or_404()
        )
        student = submission.student
        if not student or not student.user_id:
            abort(404)
        assignment_tasks = list(getattr(submission.assignment, 'tasks', None) or [])
        if not any(int(getattr(item, 'task_id', 0) or 0) == int(task_id) for item in assignment_tasks):
            abort(403)
        if not can_user_access_student(current_user, student_user_id=int(student.user_id)):
            abort(403)
        if submission.assignment and not submission.assignment.is_active:
            abort(403)
        return {
            'owner_user_id': int(student.user_id),
            'context_type': 'submission',
            'context_id': int(context_id),
            'can_write': int(current_user.id) == int(student.user_id) and not _is_parent_user(),
        }

    if context_type == 'lesson':
        if not context_id:
            abort(400, description='context_id required')
        lesson_task = (
            LessonTask.query
            .options(sa.orm.joinedload(LessonTask.lesson).joinedload(Lesson.student))
            .filter_by(lesson_task_id=int(context_id))
            .first_or_404()
        )
        lesson = lesson_task.lesson
        student = lesson.student if lesson else None
        if not student or not student.user_id:
            abort(404)
        if int(lesson_task.task_id or 0) != int(task_id):
            abort(403)
        if not can_user_access_student(current_user, student_user_id=int(student.user_id)):
            abort(403)
        return {
            'owner_user_id': int(student.user_id),
            'context_type': 'lesson',
            'context_id': int(context_id),
            'can_write': int(current_user.id) == int(student.user_id) and not _is_parent_user(),
        }

    task = Tasks.query.filter_by(task_id=int(task_id)).first()
    if not task:
        abort(404)
    return {
        'owner_user_id': int(current_user.id),
        'context_type': context_type,
        'context_id': context_id,
        'can_write': not _is_parent_user(),
    }


def _require_workspace_write(scope: dict) -> None:
    if not scope.get('can_write'):
        abort(403)


def _ensure_workspace_tables() -> None:
    """Fast non-blocking check that workspace tables exist. No DDL at request time."""
    global _workspace_tables_ok
    if _workspace_tables_ok:
        return
    if _workspace_tables_ok is False:
        abort(503, description='Workspace tables not available. Run: flask db upgrade')
    try:
        db.session.execute(sa.text('SELECT 1 FROM "StudentWorkspaceFiles" LIMIT 0'))
        db.session.execute(sa.text('SELECT 1 FROM "TaskCanvasDrawings" LIMIT 0'))
        db.session.rollback()
        _workspace_tables_ok = True
    except Exception:
        db.session.rollback()
        _workspace_tables_ok = False
        logger.error("Workspace tables missing. Run: docker compose exec web_prod flask db upgrade")
        abort(503, description='Workspace tables not available. Run: flask db upgrade')


def _ext(filename: str) -> str:
    return filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''


def _is_text(filename: str) -> bool:
    return _ext(filename) in TEXT_EXTENSIONS


def _is_excel(filename: str) -> bool:
    return _ext(filename) in SPREADSHEET_EXTENSIONS


def _write_workspace_bytes(data: bytes, user_id: int, task_id: int, filename: str) -> str:
    """Write raw bytes to workspace storage, bypassing FileStorage.save()."""
    from datetime import datetime
    key = f'workspace/{user_id}/{task_id}/{datetime.utcnow().strftime("%Y/%m")}/{filename}'
    if storage.use_s3:
        storage.client.upload_fileobj(io.BytesIO(data), storage.bucket, key)
        return f's3://{storage.bucket}/{key}'
    path = os.path.join(storage.local_upload_dir, key.replace('/', os.sep))
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'wb') as f:
        f.write(data)
    return key


def _resolve_local_path(storage_path: str) -> str | None:
    """Resolve a StorageService key to an absolute local path, if local storage is used."""
    if storage_path.startswith('s3://'):
        return None
    base = storage.local_upload_dir
    if not base:
        return None
    full = os.path.join(base, storage_path.replace('/', os.sep))
    return full if os.path.isfile(full) else None


def _read_file_bytes(ws_file: StudentWorkspaceFile) -> bytes | None:
    """Read file bytes from storage (local or S3)."""
    local = _resolve_local_path(ws_file.storage_path)
    if local:
        with open(local, 'rb') as f:
            return f.read()
    if storage.use_s3 and ws_file.storage_path.startswith('s3://'):
        s3_key = ws_file.storage_path.split('/', 3)[-1]
        resp = storage.client.get_object(Bucket=storage.bucket, Key=s3_key)
        return resp['Body'].read()
    return None


def _parse_spreadsheet(data: bytes, filename: str | None = None) -> list[dict]:
    """Parse spreadsheet files: .xlsx/.xlsm via openpyxl, .xls via xlrd, .ods via odfpy."""
    ext = _ext(filename or '')
    sheets: list[dict] = []

    if ext == 'xls':
        import xlrd
        wb = xlrd.open_workbook(file_contents=data)
        for sheet in wb.sheets():
            rows = []
            for row_idx in range(sheet.nrows):
                row = []
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell_value(row_idx, col_idx)
                    row.append('' if cell is None else str(cell))
                rows.append(row)
            sheets.append({'name': sheet.name, 'rows': rows})
        return sheets

    if ext == 'ods':
        from odf.opendocument import load as odf_load
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
        doc = odf_load(io.BytesIO(data))
        for table in doc.getElementsByType(Table):
            tbl_name = table.getAttribute('name') or 'Sheet'
            rows = []
            for tr in table.getElementsByType(TableRow):
                row: list[str] = []
                for tc in tr.getElementsByType(TableCell):
                    repeat = int(tc.getAttribute('numbercolumnsrepeated') or 1)
                    parts = []
                    for p in tc.getElementsByType(P):
                        text_parts = []
                        for node in p.childNodes:
                            val = str(node) if hasattr(node, '__str__') else ''
                            text_parts.append(val)
                        parts.append(''.join(text_parts))
                    cell_val = '\n'.join(parts)
                    for _ in range(min(repeat, 500)):
                        row.append(cell_val)
                if any(c != '' for c in row):
                    rows.append(row)
            sheets.append({'name': tbl_name, 'rows': rows})
        return sheets

    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([
                str(cell) if cell is not None else ''
                for cell in row
            ])
        sheets.append({'name': ws.title, 'rows': rows})
    wb.close()
    return sheets


def _normalize_task_file_url(raw_url: str) -> str:
    """Normalize task file URL to an absolute URL."""
    url = (raw_url or '').strip()
    if not url:
        return ''
    if url.startswith('//'):
        return 'https:' + url
    if url.startswith('/'):
        return 'https://kompege.ru' + url
    return url


def _attachment_root_dirs() -> list[str]:
    """Collect existing task-attachment root directories."""
    root_candidates: list[str] = []
    custom_root = current_app.config.get('TASK_ATTACHMENTS_ROOT')
    if custom_root:
        root_candidates.append(custom_root)
    root_candidates.append(os.path.join(current_app.root_path, 'uploads', 'task_attachments'))
    root_candidates.append(os.path.join(current_app.root_path, '..', 'uploads', 'task_attachments'))
    if storage.local_upload_dir:
        root_candidates.append(os.path.join(storage.local_upload_dir, 'task_attachments'))

    seen: set[str] = set()
    result: list[str] = []
    for root in root_candidates:
        if root and root not in seen and os.path.isdir(root):
            seen.add(root)
            result.append(root)
    return result


def _read_task_attachment_from_local_path(task_id: int, file_path: str, file_name: str | None) -> bytes | None:
    """
    Read task attachment bytes from local storage.
    Supports:
      - '/attachments/task/<task_id>/<filename>' paths
      - direct relative paths under TASK_ATTACHMENTS_ROOT
      - fallback '<TASK_ATTACHMENTS_ROOT>/<task_id>/<filename>'
      - file_name-only lookup when file_path is empty
    """
    attachment_roots = _attachment_root_dirs()
    if not attachment_roots:
        return None

    normalized_path = (file_path or '').strip()

    for att_root in attachment_roots:
        candidates: list[str] = []

        if normalized_path:
            path_only = normalized_path.split('?', 1)[0].strip()
            if path_only.startswith('/attachments/task/'):
                parts = [p for p in path_only.split('/') if p]
                if len(parts) >= 4:
                    rel_parts = parts[2:]
                    candidates.append(os.path.join(att_root, *rel_parts))
            elif path_only.startswith('/'):
                candidates.append(os.path.join(att_root, path_only.lstrip('/')))
            else:
                candidates.append(os.path.join(att_root, path_only.replace('/', os.sep)))

        if file_name:
            candidates.append(os.path.join(att_root, str(task_id), file_name))

        for local_path in candidates:
            try:
                if os.path.isfile(local_path):
                    with open(local_path, 'rb') as f:
                        return f.read()
            except Exception:
                continue
    return None


def _read_task_attachment_bytes(task_id: int, file_path: str | None, file_url: str | None, file_name: str | None) -> bytes | None:
    """Read attachment bytes: local disk first (even without file_path), then HTTP URL."""
    local_bytes = _read_task_attachment_from_local_path(task_id, file_path or '', file_name)
    if local_bytes is not None:
        logger.info("task attachment read from local disk: task_id=%s name=%s", task_id, file_name)
        return local_bytes

    if file_url:
        url = _normalize_task_file_url(file_url)
        if not url:
            return None
        logger.info("task attachment fetching from URL: task_id=%s url=%s", task_id, url)
        try:
            resp = http_requests.get(url, timeout=(5, 20))
            resp.raise_for_status()
            return resp.content
        except Exception as e:
            logger.warning("Failed to download task file %s: %s", url, e)
            return None
    return None


def _append_download_name(url: str, file_name: str | None) -> str:
    if not url or not file_name:
        return url
    sep = '&' if '?' in url else '?'
    return f"{url}{sep}download_name={file_name}"


def _build_task_download_url(task_id: int, file_path: str | None, file_url: str | None, file_name: str | None) -> str | None:
    if file_path:
        path = (file_path or '').strip()
        if path.startswith('/'):
            return _append_download_name(path, file_name)
        safe_name = (file_name or path.split('/')[-1].split('?')[0] or 'file').strip()
        return _append_download_name(f"/attachments/task/{task_id}/{safe_name}", safe_name)
    if file_url:
        normalized = _normalize_task_file_url(file_url)
        if not normalized:
            return None
        return normalized
    return None


def _is_text_previewable(filename: str, mime_type: str | None = None) -> bool:
    ext = _ext(filename)
    if ext in TEXT_EXTENSIONS:
        return True
    return bool(mime_type and mime_type.startswith('text/'))


# ---------------------------------------------------------------------------
#  File workspace endpoints
# ---------------------------------------------------------------------------

@workspace_bp.route('/workspace/files', methods=['GET'])
@login_required
def list_files():
    _ensure_workspace_tables()
    task_id = request.args.get('task_id', type=int)
    context_type = request.args.get('context_type', 'submission')
    context_id = request.args.get('context_id', type=int)
    if not task_id:
        return jsonify({'success': False, 'error': 'task_id required'}), 400
    scope = _workspace_context_scope(task_id, context_type, context_id)
    files = StudentWorkspaceFile.query.filter_by(
        user_id=scope['owner_user_id'],
        task_id=task_id,
        context_type=scope['context_type'],
    )
    if scope['context_id'] is not None:
        files = files.filter_by(context_id=scope['context_id'])
    files = files.order_by(StudentWorkspaceFile.created_at.desc()).all()
    return jsonify({'success': True, 'files': [f.to_dict() for f in files]})


@workspace_bp.route('/workspace/copy-from-task', methods=['POST'])
@login_required
@limiter.limit('30/minute')
def copy_from_task():
    """Copy a file from task's attached_files into student's workspace."""
    _ensure_workspace_tables()
    data = request.get_json(silent=True) or {}
    task_id = data.get('task_id')
    file_index = data.get('file_index')
    context_type = data.get('context_type', 'submission')
    context_id = data.get('context_id')

    if task_id is None or file_index is None:
        return jsonify({'success': False, 'error': 'task_id and file_index required'}), 400
    try:
        file_index = int(file_index)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'file_index must be integer'}), 400
    scope = _workspace_context_scope(int(task_id), context_type, context_id)
    _require_workspace_write(scope)

    existing_count = StudentWorkspaceFile.query.filter_by(
        user_id=scope['owner_user_id'], task_id=task_id, context_type=scope['context_type'],
        context_id=scope['context_id'],
    ).count()
    if existing_count >= MAX_FILES_PER_TASK:
        return jsonify({'success': False, 'error': f'Максимум {MAX_FILES_PER_TASK} файлов'}), 400

    task = Tasks.query.filter_by(task_id=task_id).first()
    if not task or not task.attached_files:
        return jsonify({'success': False, 'error': 'Задание или файлы не найдены'}), 404

    try:
        files_list = json.loads(task.attached_files) if isinstance(task.attached_files, str) else task.attached_files
    except Exception:
        return jsonify({'success': False, 'error': 'Не удалось распарсить файлы задания'}), 500

    if not isinstance(files_list, list) or file_index >= len(files_list):
        return jsonify({'success': False, 'error': 'Файл с таким индексом не найден'}), 404

    file_info = files_list[file_index]
    file_url = None
    file_name = None
    file_path = None

    if isinstance(file_info, str):
        file_url = file_info
    elif isinstance(file_info, dict):
        file_path = file_info.get('path')
        file_url = file_info.get('url')
        file_name = file_info.get('name') or file_info.get('filename')

    if not file_name:
        raw = file_path or file_url or 'file'
        file_name = raw.split('/')[-1].split('?')[0] or 'file'

    file_bytes = _read_task_attachment_bytes(task_id=task_id, file_path=file_path, file_url=file_url, file_name=file_name)
    if file_bytes is None:
        return jsonify({'success': False, 'error': 'Не удалось скачать файл'}), 502

    safe_name = secure_filename(file_name) or 'file'

    try:
        storage_key = _write_workspace_bytes(
            file_bytes, scope['owner_user_id'], task_id, safe_name,
        )
    except Exception as e:
        logger.error(f"Storage upload failed: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Ошибка сохранения файла'}), 500

    mime = mimetypes.guess_type(safe_name)[0] or 'application/octet-stream'
    ws_file = StudentWorkspaceFile(
        user_id=scope['owner_user_id'],
        task_id=task_id,
        context_type=scope['context_type'],
        context_id=scope['context_id'],
        original_filename=safe_name,
        current_filename=safe_name,
        storage_path=storage_key,
        file_size=len(file_bytes),
        mime_type=mime,
        is_from_task=True,
    )
    db.session.add(ws_file)
    db.session.commit()
    return jsonify({'success': True, 'file': ws_file.to_dict()})


@workspace_bp.route('/workspace/upload', methods=['POST'])
@login_required
@limiter.limit('30/minute')
def upload_file():
    """Upload a file to the student's workspace."""
    _ensure_workspace_tables()
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Файл не найден'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'success': False, 'error': 'Файл не выбран'}), 400

    task_id = request.form.get('task_id', type=int)
    context_type = request.form.get('context_type', 'submission')
    context_id = request.form.get('context_id', type=int)
    if not task_id:
        return jsonify({'success': False, 'error': 'task_id required'}), 400
    scope = _workspace_context_scope(task_id, context_type, context_id)
    _require_workspace_write(scope)

    ext = _ext(f.filename)
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({'success': False, 'error': f'Недопустимый формат: .{ext}'}), 400

    f.stream.seek(0, 2)
    size = f.stream.tell()
    f.stream.seek(0)
    
    # Check overall size across all files for this task
    current_files = StudentWorkspaceFile.query.filter_by(
        user_id=scope['owner_user_id'], task_id=task_id, context_type=scope['context_type'],
        context_id=scope['context_id'],
    ).all()
    
    total_size = sum(file.file_size for file in current_files if file.file_size) + size
    if total_size > 50 * 1024 * 1024: # 50 MB total limit per task workspace
        return jsonify({'success': False, 'error': 'Превышен общий лимит (50 МБ) для файлов задания'}), 400

    if size > MAX_WORKSPACE_FILE_SIZE:
        return jsonify({'success': False, 'error': 'Файл слишком большой (макс. 10 МБ)'}), 400

    if len(current_files) >= MAX_FILES_PER_TASK:
        return jsonify({'success': False, 'error': f'Максимум {MAX_FILES_PER_TASK} файлов'}), 400

    safe_name = secure_filename(f.filename) or 'file'
    try:
        storage_key = storage.upload_file(
            f, folder=f'workspace/{scope["owner_user_id"]}/{task_id}', filename=safe_name,
        )
    except Exception as e:
        logger.error(f"Workspace upload failed: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Ошибка загрузки'}), 500

    mime = mimetypes.guess_type(safe_name)[0] or 'application/octet-stream'
    ws_file = StudentWorkspaceFile(
        user_id=scope['owner_user_id'],
        task_id=task_id,
        context_type=scope['context_type'],
        context_id=scope['context_id'],
        original_filename=safe_name,
        current_filename=safe_name,
        storage_path=storage_key,
        file_size=size,
        mime_type=mime,
        is_from_task=False,
    )
    db.session.add(ws_file)
    db.session.commit()
    return jsonify({'success': True, 'file': ws_file.to_dict()})


@workspace_bp.route('/workspace/<int:file_id>/rename', methods=['POST'])
@login_required
def rename_file(file_id):
    _ensure_workspace_tables()
    ws_file = StudentWorkspaceFile.query.get_or_404(file_id)
    scope = _workspace_context_scope(ws_file.task_id, ws_file.context_type, ws_file.context_id)
    _require_workspace_write(scope)
    if ws_file.user_id != scope['owner_user_id']:
        abort(403)
    data = request.get_json(silent=True) or {}
    new_name = (data.get('new_name') or '').strip()
    if not new_name or len(new_name) > 255:
        return jsonify({'success': False, 'error': 'Некорректное имя файла'}), 400
    if re.search(r'[/\\<>:"|?*\x00-\x1f]', new_name):
        return jsonify({'success': False, 'error': 'Имя содержит недопустимые символы'}), 400
    ws_file.current_filename = new_name
    db.session.commit()
    return jsonify({'success': True, 'file': ws_file.to_dict()})


@workspace_bp.route('/workspace/<int:file_id>', methods=['DELETE'])
@login_required
def delete_file(file_id):
    _ensure_workspace_tables()
    ws_file = StudentWorkspaceFile.query.get_or_404(file_id)
    scope = _workspace_context_scope(ws_file.task_id, ws_file.context_type, ws_file.context_id)
    _require_workspace_write(scope)
    if ws_file.user_id != scope['owner_user_id']:
        abort(403)
    try:
        storage.delete_file(ws_file.storage_path)
    except Exception as e:
        logger.warning(f"Failed to delete storage file {ws_file.storage_path}: {e}")
    db.session.delete(ws_file)
    db.session.commit()
    return jsonify({'success': True})


@workspace_bp.route('/workspace/<int:file_id>/content', methods=['GET'])
@login_required
def file_content(file_id):
    """Return file content: text as plain text, Excel as JSON with sheets."""
    _ensure_workspace_tables()
    ws_file = StudentWorkspaceFile.query.get_or_404(file_id)
    scope = _workspace_context_scope(ws_file.task_id, ws_file.context_type, ws_file.context_id)
    if ws_file.user_id != scope['owner_user_id']:
        abort(403)

    data = _read_file_bytes(ws_file)
    if data is None:
        return jsonify({'success': False, 'error': 'Файл не найден в хранилище'}), 404

    fname = ws_file.current_filename
    ext = _ext(fname)

    if ext in SPREADSHEET_EXTENSIONS:
        try:
            sheets = _parse_spreadsheet(data, fname)
            return jsonify({
                'success': True,
                'type': 'excel',
                'filename': fname,
                'sheets': sheets,
            })
        except Exception as e:
            logger.error(f"Excel parse error for {fname}: {e}", exc_info=True)
            return jsonify({'success': False, 'error': 'Не удалось прочитать Excel-файл'}), 500

    if not _is_text_previewable(fname, ws_file.mime_type):
        return jsonify({
            'success': True,
            'type': 'unsupported',
            'filename': fname,
            'error': 'Предпросмотр для этого формата пока не поддерживается',
            'download_url': url_for('workspace.download_file', file_id=file_id),
        })

    mode = CODEMIRROR_MODES.get(ext, 'text/plain')
    try:
        text = data.decode('utf-8')
    except UnicodeDecodeError:
        try:
            text = data.decode('cp1251')
        except UnicodeDecodeError:
            text = data.decode('latin-1')

    return jsonify({
        'success': True,
        'type': 'text',
        'filename': fname,
        'content': text,
        'mode': mode,
    })


@workspace_bp.route('/workspace/<int:file_id>/download', methods=['GET'])
@login_required
def download_file(file_id):
    _ensure_workspace_tables()
    ws_file = StudentWorkspaceFile.query.get_or_404(file_id)
    scope = _workspace_context_scope(ws_file.task_id, ws_file.context_type, ws_file.context_id)
    if ws_file.user_id != scope['owner_user_id']:
        abort(403)

    local = _resolve_local_path(ws_file.storage_path)
    if local:
        return send_file(local, as_attachment=True, download_name=ws_file.current_filename)

    if storage.use_s3:
        url = storage.get_url(ws_file.storage_path)
        if url:
            from flask import redirect
            return redirect(url)

    abort(404)


@workspace_bp.route('/workspace/create', methods=['POST'])
@login_required
@limiter.limit('30/minute')
def create_file():
    """Create a new empty file in the student's workspace."""
    _ensure_workspace_tables()
    data = request.get_json(silent=True) or {}
    task_id = data.get('task_id')
    filename = (data.get('filename') or '').strip()
    context_type = data.get('context_type', 'submission')
    context_id = data.get('context_id')

    if not task_id or not filename:
        return jsonify({'success': False, 'error': 'task_id и filename обязательны'}), 400
    scope = _workspace_context_scope(int(task_id), context_type, context_id)
    _require_workspace_write(scope)
    if len(filename) > 255 or re.search(r'[/\\<>:"|?*\x00-\x1f]', filename):
        return jsonify({'success': False, 'error': 'Некорректное имя файла'}), 400

    ext = _ext(filename)
    if ext and ext not in ALLOWED_EXTENSIONS:
        return jsonify({'success': False, 'error': f'Недопустимый формат: .{ext}'}), 400

    existing_count = StudentWorkspaceFile.query.filter_by(
        user_id=scope['owner_user_id'], task_id=task_id, context_type=scope['context_type'],
        context_id=scope['context_id'],
    ).count()
    if existing_count >= MAX_FILES_PER_TASK:
        return jsonify({'success': False, 'error': f'Максимум {MAX_FILES_PER_TASK} файлов'}), 400

    initial_content = (data.get('content') or '').encode('utf-8')
    safe_name = secure_filename(filename) or 'file.txt'

    try:
        storage_key = _write_workspace_bytes(
            initial_content, scope['owner_user_id'], task_id, safe_name,
        )
    except Exception as e:
        logger.error("Workspace create failed: %s", e, exc_info=True)
        return jsonify({'success': False, 'error': 'Ошибка создания файла'}), 500

    mime = mimetypes.guess_type(safe_name)[0] or 'text/plain'
    ws_file = StudentWorkspaceFile(
        user_id=scope['owner_user_id'],
        task_id=task_id,
        context_type=scope['context_type'],
        context_id=scope['context_id'],
        original_filename=safe_name,
        current_filename=safe_name,
        storage_path=storage_key,
        file_size=len(initial_content),
        mime_type=mime,
        is_from_task=False,
    )
    db.session.add(ws_file)
    db.session.commit()
    return jsonify({'success': True, 'file': ws_file.to_dict()})


@workspace_bp.route('/workspace/<int:file_id>/save-content', methods=['POST'])
@login_required
@limiter.limit('60/minute')
def save_content(file_id):
    """Save updated text content for a workspace file."""
    _ensure_workspace_tables()
    ws_file = StudentWorkspaceFile.query.get_or_404(file_id)
    scope = _workspace_context_scope(ws_file.task_id, ws_file.context_type, ws_file.context_id)
    _require_workspace_write(scope)
    if ws_file.user_id != scope['owner_user_id']:
        abort(403)

    data = request.get_json(silent=True) or {}
    content = data.get('content')
    if content is None:
        return jsonify({'success': False, 'error': 'content обязателен'}), 400

    content_bytes = content.encode('utf-8')
    if len(content_bytes) > MAX_WORKSPACE_FILE_SIZE:
        return jsonify({'success': False, 'error': 'Файл слишком большой (макс. 10 МБ)'}), 400

    try:
        new_key = _write_workspace_bytes(
            content_bytes, scope['owner_user_id'], ws_file.task_id,
            secure_filename(ws_file.current_filename) or 'file',
        )
    except Exception as e:
        logger.error("save-content upload failed: %s", e, exc_info=True)
        return jsonify({'success': False, 'error': 'Ошибка сохранения'}), 500

    try:
        if ws_file.storage_path != new_key:
            storage.delete_file(ws_file.storage_path)
    except Exception:
        pass

    ws_file.storage_path = new_key
    ws_file.file_size = len(content_bytes)
    db.session.commit()
    return jsonify({'success': True, 'file': ws_file.to_dict()})


# ---------------------------------------------------------------------------
#  Task attachment direct preview (without copying to workspace first)
# ---------------------------------------------------------------------------

@workspace_bp.route('/workspace/task-file-content', methods=['GET'])
@login_required
def task_file_content():
    """Preview a task's attached file directly without copying it to workspace."""
    task_id = request.args.get('task_id', type=int)
    file_index = request.args.get('file_index', type=int)
    logger.info("[task-file-content] START task_id=%s file_index=%s user=%s", task_id, file_index, current_user.id)

    if task_id is None or file_index is None:
        return jsonify({'success': False, 'error': 'task_id and file_index required'}), 400
    try:
        file_index = int(file_index)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'file_index must be integer'}), 400
    context_type = request.args.get('context_type', 'submission')
    context_id = request.args.get('context_id', type=int)
    _workspace_context_scope(task_id, context_type, context_id)

    task = Tasks.query.filter_by(task_id=task_id).first()
    if not task or not task.attached_files:
        logger.warning("[task-file-content] task not found or no attachments: task_id=%s", task_id)
        return jsonify({'success': False, 'error': 'Задание не найдено'}), 404

    try:
        files_list = json.loads(task.attached_files) if isinstance(task.attached_files, str) else task.attached_files
    except Exception:
        return jsonify({'success': False, 'error': 'Ошибка чтения файлов задания'}), 500

    if not isinstance(files_list, list) or file_index >= len(files_list):
        return jsonify({'success': False, 'error': 'Файл не найден'}), 404

    file_info = files_list[file_index]
    file_url = None
    file_name = None
    file_path = None

    if isinstance(file_info, str):
        file_url = file_info
    elif isinstance(file_info, dict):
        file_path = file_info.get('path')
        file_url = file_info.get('url')
        file_name = file_info.get('name') or file_info.get('filename')

    if not file_name:
        raw = file_path or file_url or 'file'
        file_name = raw.split('/')[-1].split('?')[0] or 'file'

    ext = _ext(file_name)
    logger.info("[task-file-content] resolved file_name=%s ext=%s path=%s url=%s", file_name, ext, file_path, file_url)

    download_url = _build_task_download_url(
        task_id=task_id,
        file_path=file_path,
        file_url=file_url,
        file_name=file_name,
    )
    if ext not in SPREADSHEET_EXTENSIONS and not _is_text_previewable(file_name):
        return jsonify({
            'success': True,
            'type': 'unsupported',
            'filename': file_name,
            'error': 'Предпросмотр для этого формата пока не поддерживается',
            'download_url': download_url,
        })

    logger.info("[task-file-content] reading attachment bytes ...")
    file_bytes = _read_task_attachment_bytes(task_id=task_id, file_path=file_path, file_url=file_url, file_name=file_name)
    if file_bytes is None:
        logger.warning("[task-file-content] could not read file bytes for task_id=%s name=%s", task_id, file_name)
        return jsonify({
            'success': True,
            'type': 'unsupported',
            'filename': file_name,
            'error': 'Файл не найден ни локально, ни по URL. Попробуйте скачать.',
            'download_url': download_url,
        })

    logger.info("[task-file-content] got %d bytes, processing as %s", len(file_bytes), ext)

    if ext in SPREADSHEET_EXTENSIONS:
        try:
            sheets = _parse_spreadsheet(file_bytes, file_name)
            return jsonify({
                'success': True, 'type': 'excel',
                'filename': file_name, 'sheets': sheets,
            })
        except Exception as e:
            logger.error("Excel parse error: %s", e, exc_info=True)
            return jsonify({'success': False, 'error': 'Не удалось прочитать Excel'}), 500

    if not _is_text_previewable(file_name):
        return jsonify({
            'success': True,
            'type': 'unsupported',
            'filename': file_name,
            'error': 'Предпросмотр для этого формата пока не поддерживается',
            'download_url': download_url,
        })

    mode = CODEMIRROR_MODES.get(ext, 'text/plain')
    try:
        text = file_bytes.decode('utf-8')
    except UnicodeDecodeError:
        try:
            text = file_bytes.decode('cp1251')
        except UnicodeDecodeError:
            text = file_bytes.decode('latin-1')

    return jsonify({
        'success': True, 'type': 'text',
        'filename': file_name, 'content': text, 'mode': mode,
    })


# ---------------------------------------------------------------------------
#  Canvas drawing endpoints
# ---------------------------------------------------------------------------

@workspace_bp.route('/api/canvas/save', methods=['POST'])
@login_required
@limiter.limit('60/minute')
def canvas_save():
    _ensure_workspace_tables()
    data = request.get_json(silent=True) or {}
    task_id = data.get('task_id')
    if not task_id:
        return jsonify({'success': False, 'error': 'task_id required'}), 400
    context_type = data.get('context_type', 'submission')
    context_id = data.get('context_id')
    scope = _workspace_context_scope(int(task_id), context_type, context_id)
    _require_workspace_write(scope)
    strokes = data.get('strokes', '[]')
    if isinstance(strokes, (list, dict)):
        strokes = json.dumps(strokes)

    thumbnail = data.get('thumbnail')

    drawing = TaskCanvasDrawing.query.filter_by(
        user_id=scope['owner_user_id'],
        task_id=task_id,
        context_type=scope['context_type'],
        context_id=scope['context_id'],
    ).first()

    if drawing:
        drawing.strokes_json = strokes
        if thumbnail:
            drawing.thumbnail_url = thumbnail
    else:
        drawing = TaskCanvasDrawing(
            user_id=scope['owner_user_id'],
            task_id=task_id,
            context_type=scope['context_type'],
            context_id=scope['context_id'],
            strokes_json=strokes,
            thumbnail_url=thumbnail,
        )
        db.session.add(drawing)

    db.session.commit()
    return jsonify({'success': True, 'id': drawing.id})


@workspace_bp.route('/api/canvas/load', methods=['GET'])
@login_required
@limiter.limit('30/minute')
def canvas_load():
    _ensure_workspace_tables()
    task_id = request.args.get('task_id', type=int)
    if not task_id:
        return jsonify({'success': False, 'error': 'task_id required'}), 400
    context_type = request.args.get('context_type', 'submission')
    context_id = request.args.get('context_id', type=int)
    scope = _workspace_context_scope(task_id, context_type, context_id)

    drawing = TaskCanvasDrawing.query.filter_by(
        user_id=scope['owner_user_id'],
        task_id=task_id,
        context_type=scope['context_type'],
        context_id=scope['context_id'],
    ).first()

    if not drawing:
        return jsonify({'success': True, 'strokes': '[]', 'exists': False})

    return jsonify({
        'success': True,
        'exists': True,
        'strokes': drawing.strokes_json,
        'id': drawing.id,
        'updated_at': drawing.updated_at.isoformat() if drawing.updated_at else None,
    })


@workspace_bp.route('/api/canvas/view/<int:target_user_id>', methods=['GET'])
@login_required
def canvas_view(target_user_id):
    """Teacher views student's canvas or user views own canvas."""
    _ensure_workspace_tables()
    task_id = request.args.get('task_id', type=int)
    if not task_id:
        return jsonify({'success': False, 'error': 'task_id required'}), 400
    context_type = request.args.get('context_type', 'submission')
    context_id = request.args.get('context_id', type=int)
    scope = _workspace_context_scope(task_id, context_type, context_id)
    if int(target_user_id) != int(scope['owner_user_id']):
        abort(403)

    drawing = TaskCanvasDrawing.query.filter_by(
        user_id=scope['owner_user_id'],
        task_id=task_id,
        context_type=scope['context_type'],
        context_id=scope['context_id'],
    ).first()

    if not drawing:
        return jsonify({'success': True, 'exists': False, 'strokes': '[]'})

    return jsonify({
        'success': True,
        'exists': True,
        'strokes': drawing.strokes_json,
        'thumbnail_url': drawing.thumbnail_url,
        'updated_at': drawing.updated_at.isoformat() if drawing.updated_at else None,
    })


@workspace_bp.route('/api/canvas/list', methods=['GET'])
@login_required
def canvas_list():
    """Teacher lists all canvases for a student (optionally filtered by task)."""
    _ensure_workspace_tables()
    student_user_id = request.args.get('student_user_id', type=int)
    task_id = request.args.get('task_id', type=int)
    if not student_user_id:
        return jsonify({'success': False, 'error': 'student_user_id required'}), 400
    if not can_user_access_student(current_user, student_user_id=student_user_id):
        abort(403)

    q = TaskCanvasDrawing.query.filter_by(user_id=student_user_id)
    if task_id:
        q = q.filter_by(task_id=task_id)
    drawings = q.order_by(TaskCanvasDrawing.updated_at.desc()).limit(50).all()

    return jsonify({
        'success': True,
        'drawings': [d.to_dict() for d in drawings],
    })
